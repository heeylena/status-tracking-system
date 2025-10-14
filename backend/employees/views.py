"""
API Views for Employee Status Tracking System.
"""
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.utils import timezone
from django.db.models import Sum, Count, Q
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

from .models import Employee, Status, StatusLog
from .serializers import (
    EmployeeListSerializer,
    EmployeeDetailSerializer,
    StatusSerializer,
    StatusLogSerializer,
    ChangeStatusSerializer,
    EmployeeStatisticsSerializer,
)


class EmployeeViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Employee operations.
    """
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        """Get active employees only by default."""
        queryset = Employee.objects.filter(is_active=True)
        return queryset.select_related().prefetch_related('status_logs__status')
    
    def get_serializer_class(self):
        """Return appropriate serializer based on action."""
        if self.action == 'list':
            return EmployeeListSerializer
        return EmployeeDetailSerializer
    
    @action(detail=True, methods=['post'])
    def change_status(self, request, pk=None):
        """
        Change employee status with proper overdue calculation.
        
        Business logic:
        1. Get current active StatusLog
        2. If planned_end_time exists and is in past, calculate overdue_duration
        3. Set end_time on current log
        4. Create new StatusLog with new status
        """
        employee = self.get_object()
        serializer = ChangeStatusSerializer(data=request.data)
        
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        validated_data = serializer.validated_data
        new_status = Status.objects.get(id=validated_data['status_id'])
        
        # Close current status log if exists
        current_log = employee.get_current_status_log()
        if current_log:
            current_log.end_time = timezone.now()
            current_log.calculate_and_save_overdue_duration()
            current_log.save()
        
        # Create new status log
        new_log = StatusLog.objects.create(
            employee=employee,
            status=new_status,
            planned_end_time=validated_data.get('planned_end_time'),
            notes=validated_data.get('notes', ''),
            created_by=request.user
        )
        
        # Return updated employee data
        employee_serializer = EmployeeDetailSerializer(employee)
        return Response(employee_serializer.data, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=['get'])
    def history(self, request, pk=None):
        """
        Get status history for an employee with pagination.
        """
        employee = self.get_object()
        logs = StatusLog.objects.filter(employee=employee).select_related('status', 'created_by')
        
        # Apply pagination
        page = self.paginate_queryset(logs)
        if page is not None:
            serializer = StatusLogSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = StatusLogSerializer(logs, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def statistics(self, request, pk=None):
        """
        Get aggregated time statistics for an employee.
        Shows total time spent in each status.
        """
        employee = self.get_object()
        
        # Aggregate statistics by status
        stats = []
        statuses = Status.objects.all()
        
        for s in statuses:
            logs = StatusLog.objects.filter(employee=employee, status=s)
            
            total_seconds = 0
            total_overdue = 0
            
            for log in logs:
                if log.end_time:
                    duration = (log.end_time - log.start_time).total_seconds()
                else:
                    duration = (timezone.now() - log.start_time).total_seconds()
                
                total_seconds += int(duration)
                total_overdue += log.overdue_duration
            
            if total_seconds > 0:  # Only include statuses with time logged
                stats.append({
                    'status_name': s.name,
                    'status_color': s.color,
                    'total_seconds': total_seconds,
                    'count': logs.count(),
                    'total_overdue_seconds': total_overdue,
                })
        
        serializer = EmployeeStatisticsSerializer(stats, many=True)
        return Response(serializer.data)


class StatusViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Status operations.
    """
    permission_classes = [IsAuthenticated]
    serializer_class = StatusSerializer
    queryset = Status.objects.filter(is_active=True).order_by('display_order', 'name')


class ReportViewSet(viewsets.ViewSet):
    """
    ViewSet for generating reports.
    """
    permission_classes = [IsAuthenticated]
    
    @action(detail=False, methods=['get', 'post'])
    def excel(self, request):
        """
        Generate Excel report of employee statuses.
        
        GET: Download all current employee statuses
        POST: Download custom report with filters (employee_id, status_id, start_date, end_date)
        """
        # Parse filters from request
        employee_id = request.data.get('employee_id') if request.method == 'POST' else None
        status_id = request.data.get('status_id') if request.method == 'POST' else None
        start_date = request.data.get('start_date') if request.method == 'POST' else None
        end_date = request.data.get('end_date') if request.method == 'POST' else None
        
        # Build query
        logs = StatusLog.objects.select_related('employee', 'status', 'created_by')
        
        if employee_id:
            logs = logs.filter(employee_id=employee_id)
        if status_id:
            logs = logs.filter(status_id=status_id)
        if start_date:
            logs = logs.filter(start_time__gte=start_date)
        if end_date:
            logs = logs.filter(start_time__lte=end_date)
        
        logs = logs.order_by('-start_time')
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Status Report"
        
        # Define headers
        headers = [
            'Employee', 'Status', 'Start Time', 'End Time',
            'Planned End', 'Duration (hours)', 'Overdue (hours)', 'Notes'
        ]
        
        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add data rows
        for row_num, log in enumerate(logs, 2):
            ws.cell(row=row_num, column=1).value = log.employee.name
            ws.cell(row=row_num, column=2).value = log.status.name
            ws.cell(row=row_num, column=3).value = log.start_time.strftime('%Y-%m-%d %H:%M:%S')
            ws.cell(row=row_num, column=4).value = log.end_time.strftime('%Y-%m-%d %H:%M:%S') if log.end_time else 'Active'
            ws.cell(row=row_num, column=5).value = log.planned_end_time.strftime('%Y-%m-%d %H:%M:%S') if log.planned_end_time else 'N/A'
            
            # Calculate duration
            if log.end_time:
                duration_hours = (log.end_time - log.start_time).total_seconds() / 3600
            else:
                duration_hours = (timezone.now() - log.start_time).total_seconds() / 3600
            ws.cell(row=row_num, column=6).value = round(duration_hours, 2)
            
            # Overdue duration
            overdue_hours = log.overdue_duration / 3600
            ws.cell(row=row_num, column=7).value = round(overdue_hours, 2) if overdue_hours > 0 else 0
            
            ws.cell(row=row_num, column=8).value = log.notes
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Generate response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'employee_status_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        
        wb.save(response)
        return response
